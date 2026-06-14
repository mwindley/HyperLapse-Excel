#!/usr/bin/env python3
"""
HyperLapse - Pano Planner image.

A DESIGN-TIME exploration aid (run once per lens, not every shoot). Reads the
two config blocks on the PANO sheet (landscape + portrait) and draws, for each:

  - the coverage strip: every frame to scale, with its VIRGIN DISTORTED edges
    (white wedges) and the clean centre, the subject span band + framing margin
  - the overlap readout (deg + %), the 30%/50% rule-of-thumb guides
  - the time buckets (photo vs non-photo) and the FINAL VIDEO length at the
    sheet's shoot-duration + FPS - the cost of more shots

The image is the teaching surface: newbies study the edges/overlap; experienced
operators just dial overlap to ~30% (brave about post) or ~50% (not) and check
the final-video cost. The PANO sheet FORMULAS are the contract; this only SHOWS.

Reads the same inputs the sheet formulas use, recomputes independently (so the
picture is self-checking against the sheet), and reads the two extra planner
inputs (shoot duration hr, FPS) if present, else uses defaults.

Usage:
  python3 pano_planner.py HyperLapse.xlsm out.png
"""
import sys
import math
import argparse
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
from matplotlib.collections import PatchCollection
import numpy as np
from openpyxl import load_workbook

BG = "#0d1117"
CARD = "#161b22"
FRAME = "#888780"
FRAME_EDGE = "#5F5E5A"
SUBJECT = "#1D9E75"
MARGIN = "#0F6E56"
TXT = "#e6edf3"
MUTED = "#8b949e"
GOOD = "#3fb950"
WARN = "#d29922"
BAD = "#ff5470"

FF_LONG = 36.0
FF_SHORT = 24.0
POST_SHUTTER_MS = 500.0
SLEW_MIN_MS = 500.0

DEFAULT_DUR_HR = 12.0
DEFAULT_FPS = 60.0


def _named_cell(wb, name):
    """Return the value of a single-cell named range, or None."""
    try:
        dn = wb.defined_names[name]
        for sheet, ref in dn.destinations:
            ws = wb[sheet]
            return ws[ref.replace("$", "")].value
    except Exception:
        return None
    return None


def _named_row(wb, name):
    """Return the list of values across a multi-cell named range (the offsets)."""
    out = []
    try:
        dn = wb.defined_names[name]
        for sheet, ref in dn.destinations:
            ws = wb[sheet]
            for row in ws[ref.replace("$", "")]:
                for c in row:
                    out.append(c.value)
    except Exception:
        pass
    return out


def fov_deg(focal, dim):
    return math.degrees(2 * math.atan(dim / (2 * focal)))


def read_block(wb, pfx):
    """Read one config block by prefix ('L' or 'P'). Falls back gracefully."""
    nm = "pano" + pfx + "_"
    g = lambda k, d=None: _named_cell(wb, nm + k) if _named_cell(wb, nm + k) is not None else d
    cfg = {
        "lens": g("lens", "?"),
        "orient": str(g("orient", "landscape")).lower(),
        "focal": float(g("focal", 14) or 14),
        "shots": int(g("shots", 4) or 4),
        "span": float(g("span", 180) or 180),
        "edge": float(g("edge", 40) or 40),
        "tv": float(g("tv", 20) or 20),
        "slew": float(g("slew", 20) or 20),
        "settle": float(g("settle", 800) or 800),
    }
    return cfg


def compute(cfg, dur_hr, fps):
    f = cfg["focal"]
    yaw_dim = FF_LONG if cfg["orient"] == "landscape" else FF_SHORT
    hfov = fov_deg(f, yaw_dim)
    n = max(1, cfg["shots"])
    need = cfg["span"] + 2 * cfg["edge"]
    if n <= 1:
        step = 0.0
        offsets = [0.0]
    else:
        step = (need - hfov) / (n - 1)
        start = -(n - 1) * step / 2
        offsets = [start + i * step for i in range(n)]
    overlap = hfov - step
    overlap_pct = (100 * overlap / hfov) if hfov else 0
    coverage = hfov + (n - 1) * step

    photo_ms = n * cfg["tv"] * 1000
    slew = max(0.1, cfg["slew"])
    slew_ms = max(abs(offsets[0]) / slew * 1000, SLEW_MIN_MS)
    for i in range(1, n):
        slew_ms += max(abs(offsets[i] - offsets[i - 1]) / slew * 1000, SLEW_MIN_MS)
    slew_ms += max(abs(offsets[-1]) / slew * 1000, SLEW_MIN_MS)
    settle_ms = n * cfg["settle"]
    post_ms = n * POST_SHUTTER_MS
    nonphoto_ms = slew_ms + settle_ms + post_ms
    total_ms = photo_ms + nonphoto_ms

    shoot_sec = dur_hr * 3600
    panos = int(shoot_sec / (total_ms / 1000)) if total_ms > 0 else 0
    video_sec = panos / fps if fps else 0

    return dict(hfov=hfov, n=n, need=need, step=step, offsets=offsets,
                overlap=overlap, overlap_pct=overlap_pct, coverage=coverage,
                photo_s=photo_ms / 1000, nonphoto_s=nonphoto_ms / 1000,
                total_s=total_ms / 1000, panos=panos, video_s=video_sec,
                slew_s=slew_ms / 1000, settle_s=settle_ms / 1000, post_s=post_ms / 1000)


def fmt_video(sec):
    if sec >= 60:
        return f"{int(sec // 60)}m {int(round(sec % 60))}s"
    return f"{sec:.1f}s"


def draw_block(ax, cfg, r, title):
    ax.set_facecolor(CARD)
    ax.set_xticks([]); ax.set_yticks([])
    for s in ax.spines.values():
        s.set_visible(False)

    n = r["n"]; hfov = r["hfov"]; offsets = r["offsets"]
    left = offsets[0] - hfov / 2
    right = offsets[-1] + hfov / 2
    total = right - left if right > left else 1

    ax.set_xlim(left - total * 0.04, right + total * 0.04)
    rowH = 1.0; gap = 0.3
    n_rows = n
    strip_y = 0
    ax.set_ylim(-2.2, n_rows * (rowH + gap) + 1.0)

    edge_frac = 0.18
    patches = []
    for i in range(n):
        y = strip_y + i * (rowH + gap)
        x0 = offsets[i] - hfov / 2
        ax.add_patch(Rectangle((x0, y), hfov, rowH, facecolor=FRAME, alpha=0.28,
                               edgecolor=FRAME_EDGE, lw=0.5, zorder=2))
        ew = hfov * edge_frac
        # white virgin-distorted edge wedges (left + right of each frame)
        ax.add_patch(Rectangle((x0, y), ew, rowH, facecolor="white", alpha=0.5, lw=0, zorder=3))
        ax.add_patch(Rectangle((x0 + hfov - ew, y), ew, rowH, facecolor="white", alpha=0.5, lw=0, zorder=3))
        ax.text(offsets[i], y + rowH / 2, f"{offsets[i]:+.0f}", ha="center", va="center",
                fontsize=8, color=TXT, zorder=4)

    # subject span band + framing margins, below the frames
    sy = -1.6
    ax.add_patch(Rectangle((left, sy), total, 1.0, facecolor=SUBJECT, alpha=0.16, lw=0))
    sub_l = -cfg["span"] / 2; sub_r = cfg["span"] / 2
    ax.add_patch(Rectangle((sub_l, sy), cfg["span"], 1.0, facecolor=SUBJECT, alpha=0.5, lw=0))
    ax.plot([sub_l, sub_l], [sy - 0.3, sy + 1.3], color=MARGIN, lw=0.7, ls=":")
    ax.plot([sub_r, sub_r], [sy - 0.3, sy + 1.3], color=MARGIN, lw=0.7, ls=":")
    ax.text(0, sy + 0.5, f"subject {cfg['span']:.0f}\u00b0", ha="center", va="center",
            fontsize=8.5, color="white", zorder=5)
    ax.text((left + sub_l) / 2, sy + 0.5, f"{cfg['edge']:.0f}\u00b0", ha="center", va="center",
            fontsize=8, color="white")
    ax.text((sub_r + right) / 2, sy + 0.5, f"{cfg['edge']:.0f}\u00b0", ha="center", va="center",
            fontsize=8, color="white")

    # title + lens
    ax.text(left, n_rows * (rowH + gap) + 0.55, title, fontsize=12, color=TXT, fontweight="bold")
    ax.text(right, n_rows * (rowH + gap) + 0.55, f"lens: {cfg['lens']}", fontsize=9,
            color=MUTED, ha="right")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsm")
    ap.add_argument("out")
    args = ap.parse_args()

    wb = load_workbook(args.xlsm, read_only=True, data_only=True)
    dur = _named_cell(wb, "pano_dur_hr") or DEFAULT_DUR_HR
    fps = _named_cell(wb, "pano_fps") or DEFAULT_FPS
    dur = float(dur); fps = float(fps)

    blocks = []
    for pfx, label in (("L", "Landscape"), ("P", "Portrait")):
        cfg = read_block(wb, pfx)
        r = compute(cfg, dur, fps)
        blocks.append((cfg, r, label))

    fig = plt.figure(figsize=(11, 9), facecolor=BG)
    gs = fig.add_gridspec(2, 1, height_ratios=[1, 1], hspace=0.35,
                          left=0.04, right=0.97, top=0.93, bottom=0.06)

    fig.suptitle(f"Pano planner  \u00b7  shoot {dur:.0f}h @ {fps:.0f}fps",
                 color=TXT, fontsize=14, x=0.04, ha="left", y=0.975)

    for idx, (cfg, r, label) in enumerate(blocks):
        ax = fig.add_subplot(gs[idx])
        draw_block(ax, cfg, r, label)
        # summary line beneath each block
        ov_col = BAD if r["overlap"] <= 0 else (WARN if r["overlap_pct"] < 25 else GOOD)
        cov_ok = r["coverage"] >= r["need"] - 0.5
        summ = (f"FOV {r['hfov']:.0f}\u00b0  \u00b7  step {r['step']:.0f}\u00b0  \u00b7  "
                f"overlap {r['overlap']:.0f}\u00b0 ({r['overlap_pct']:.0f}%)  \u00b7  "
                f"coverage {r['coverage']:.0f}\u00b0  \u00b7  "
                f"photo {r['photo_s']:.0f}s + non-photo {r['nonphoto_s']:.0f}s = {r['total_s']:.0f}s/pano")
        ax.text(0.5, -0.07, summ, transform=ax.transAxes, ha="center", va="top",
                fontsize=9, color=ov_col if (r['overlap_pct'] < 25 or r['overlap'] <= 0) else TXT)
        video = (f"{r['panos']:,} panos  \u2192  {fmt_video(r['video_s'])} final video"
                 + ("" if cov_ok else "   COVERAGE SHORT"))
        ax.text(0.5, -0.155, video, transform=ax.transAxes, ha="center", va="top",
                fontsize=11, color=(GOOD if cov_ok else BAD), fontweight="bold")

    fig.savefig(args.out, facecolor=BG, bbox_inches="tight", dpi=110)
    svg = args.out.rsplit(".", 1)[0] + ".svg"
    fig.savefig(svg, facecolor=BG, bbox_inches="tight")
    print(f"wrote {args.out}")


if __name__ == "__main__":
    main()
