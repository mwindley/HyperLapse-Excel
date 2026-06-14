# Diagnostic: read the plan the same way the renderer does and show what the
# arch rows yield for the track-sampling guard. Run:
#   python arch_diag.py "C:\Github\HyperLapse-Excel\HyperLapse.xlsm"
import sys
from openpyxl import load_workbook
p = sys.argv[1] if len(sys.argv)>1 else r"C:\Github\HyperLapse-Excel\HyperLapse.xlsm"
wb = load_workbook(p, read_only=True, data_only=True)
# Find the plan sheet + the MIDDLE header row (look for "Step" + "Action" + "Target")
for ws in wb.worksheets:
    for hr in range(1, 12):
        vals = [str(c.value).strip() if c.value is not None else "" for c in ws[hr]]
        if "Action" in vals and "Target" in vals and "Actual (mins)" in vals:
            print(f"sheet='{ws.title}' header row={hr}")
            idx = {vals[i]: i for i in range(len(vals))}
            for rr in range(hr+1, hr+12):
                row = [c.value for c in ws[rr]]
                def g(name):
                    i = idx.get(name)
                    return row[i] if i is not None and i < len(row) else None
                act = g("Action"); tgt = g("Target"); fa = g("Fires at"); am = g("Actual (mins)")
                if act is None and tgt is None: continue
                print(f"  Action={act!r:14} Target={tgt!r:12} Fires={fa!r:10} Actual(mins)={am!r:8} type={type(am).__name__}")
            sys.exit()
print("No plan header found")
