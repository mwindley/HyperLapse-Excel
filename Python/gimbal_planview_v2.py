#!/usr/bin/env python3
"""
HyperLapse — Gimbal Plan View (#2), production renderer.
Spec: GIMBAL_PLANVIEW_BUILD.md.

Step A (resolve) + Step C (render). Reads HyperLapse.xlsm directly.

Per GP:
  - frame: earth if target in {sun,moon,gc} or Action in {Track,Track-yaw};
           else chassis.
  - cart-frame yaw/pitch accumulate (numeric Ry/Rp = absolute anchor).
  - world bearing:
        chassis : expected_cart_heading(anchor WP) + cart_frame_yaw
        earth   : object_az(fire time) + dyaw   [+ heading-correction scalar]
  - cumulative cart-frame yaw carried signed/un-wrapped (cable + midpoint).

Render:
  - cart centre, true-N up, radius = altitude (horizon rim, zenith centre)
  - GP glyph = radial line rim->endpoint, dir = world bearing,
    length proportional to pitch (endpoint sits on the altitude grid)
  - earth vs chassis styled distinctly
  - midpoint on near-180 cart-frame slews, labelled cumulative yaw
  - park-and-wait marker for below-horizon earth objects
  - shared validation flags (pitch > 80)
  - PREV/NEXT: --gp N foregrounds one GP
  - map underlay v1: --map north_up.png (decorative, clipped to dial)

Usage: python3 gimbal_planview_v2.py HyperLapse.xlsm out.png [--gp 3] [--map t.png]
"""
import sys, math, argparse, datetime
from openpyxl import load_workbook
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from matplotlib.patches import Circle, FancyArrowPatch

BG="#0d1117"; RING="#2b3340"; RINGTXT="#5b6675"; CARD="#c9d4e3"
CHASSIS="#36d399"; EARTH={"sun":"#ffb02e","gc":"#9b8cff","moon":"#7fd4ff","arch_rise":"#5ce6c0","arch_set":"#e65ca8"}
EARTH_DEF="#ffb02e"; HDG="#ff5c7a"; FLAG="#ff5470"; MID="#ffd24a"; TXT="#e6edf3"
DIM=0.22
R=1.0
NEAR180=150.0           # cart-frame |dyaw| that triggers a direction midpoint
PITCH_LIMIT=80.0
OBJ_COL={"sun":(3,4),"gc":(1,2)}   # AstroTable (az_col, alt_col) by object

def alt_to_r(alt): return max(0.0,min(1.0,(90.0-max(0,min(90,alt)))/90.0))*R
def az_to_xy(az,r):
    a=math.radians(az); return r*math.sin(a), r*math.cos(a)

# ---- GC ephemeris (ported verbatim from Astro.bas: GetGCPosition chain) ----
# Single source of truth with the cart's cubic author (AstroPush.FitAndPushTrackPath
# samples this same function). Lets the renderer draw the EXECUTED track offline.
GC_RA_DEG=266.4167; GC_DEC_DEG=-29.0078     # galactic centre J2000
def _date_to_julian(dt):
    y,m,d=dt.year,dt.month,dt.day
    hr=dt.hour+dt.minute/60.0+dt.second/3600.0
    if m<=2: y-=1; m+=12
    a=y//100; b=2-a+a//4
    return int(365.25*(y+4716))+int(30.6001*(m+1))+d+b-1524.5+hr/24.0
def _norm(deg): return deg-360.0*math.floor(deg/360.0)
def gc_az_alt(local_dt,lat,lng,utc_off):
    """GC azimuth/altitude (deg) at a LOCAL datetime. Mirrors GetGCPosition."""
    utc=local_dt-datetime.timedelta(hours=utc_off)
    n=_date_to_julian(utc)-2451545.0
    gmst=_norm(280.46061837+360.98564736629*n)
    lst=_norm(gmst+lng)
    ha=_norm(lst-GC_RA_DEG)
    if ha>180: ha-=360
    return _radec_to_altaz(ha,GC_DEC_DEG,lat)

def _radec_to_altaz(ha,dec,lat):
    har,dr,lr=map(math.radians,(ha,dec,lat))
    sa=math.sin(dr)*math.sin(lr)+math.cos(dr)*math.cos(lr)*math.cos(har)
    sa=max(-1.0,min(1.0,sa)); alt=math.degrees(math.asin(sa))
    ca=(math.sin(dr)-math.sin(math.radians(alt))*math.sin(lr))/(math.cos(math.radians(alt))*math.cos(lr))
    ca=max(-1.0,min(1.0,ca)); az=math.degrees(math.acos(ca))
    if math.sin(har)>0: az=360.0-az
    return az,alt

def _atn2(y,x): return math.degrees(math.atan2(y,x))
def sun_az_alt(local_dt,lat,lng,utc_off):
    """Sun azimuth/altitude (deg). Mirrors Astro.bas GetSunPosition."""
    utc=local_dt-datetime.timedelta(hours=utc_off)
    n=_date_to_julian(utc)-2451545.0
    L=_norm(280.46+0.9856474*n); g=_norm(357.528+0.9856003*n); gr=math.radians(g)
    lam=_norm(L+1.915*math.sin(gr)+0.02*math.sin(2*gr)); lr=math.radians(lam)
    eps=23.439-0.0000004*n; er=math.radians(eps)
    ra=_norm(_atn2(math.cos(er)*math.sin(lr),math.cos(lr)))
    dec=math.degrees(math.asin(math.sin(er)*math.sin(lr)))
    gmst=_norm(280.46061837+360.98564736629*n); lst=_norm(gmst+lng)
    ha=_norm(lst-ra)
    if ha>180: ha-=360
    return _radec_to_altaz(ha,dec,lat)

def moon_az_alt(local_dt,lat,lng,utc_off):
    """Moon azimuth/altitude (deg). Mirrors Astro.bas GetMoonPosition (Schlyter)."""
    utc=local_dt-datetime.timedelta(hours=utc_off)
    jd=_date_to_julian(utc); d=jd-2451543.5
    NN=_norm(125.1228-0.0529538083*d); ii=5.1454
    w=_norm(318.0634+0.1643573223*d); a=60.2666; e=0.0549
    m=_norm(115.3654+13.0649929509*d); mr=math.radians(m)
    E1=m+math.degrees(e*math.sin(mr)*(1+e*math.cos(mr)))
    for _ in range(10):
        E0=E1; e0r=math.radians(E0)
        E1=E0-(E0-math.degrees(e*math.sin(e0r))-m)/(1-e*math.cos(e0r))
        if abs(E1-E0)<0.001: break
    e1r=math.radians(E1)
    xv=a*(math.cos(e1r)-e); yv=a*math.sqrt(1-e*e)*math.sin(e1r)
    v=_atn2(yv,xv); r=math.hypot(xv,yv)
    NR=math.radians(NN); wv=math.radians(w+v); iR=math.radians(ii)
    xe_=r*(math.cos(NR)*math.cos(wv)-math.sin(NR)*math.sin(wv)*math.cos(iR))
    ye_=r*(math.sin(NR)*math.cos(wv)+math.cos(NR)*math.sin(wv)*math.cos(iR))
    ze_=r*math.sin(wv)*math.sin(iR)
    eclLon=_norm(_atn2(ye_,xe_)); eclLat=_atn2(ze_,math.hypot(xe_,ye_))
    ws_=_norm(282.9404+0.0000470935*d); ms=_norm(356.047+0.9856002585*d)
    Ls=_norm(ws_+ms); Lm=_norm(NN+w+m)
    Mm=m; Dm=_norm(Lm-Ls); f=_norm(Lm-NN)
    R_=math.radians
    eclLon=_norm(eclLon+(-1.274*math.sin(R_(Mm-2*Dm))+0.658*math.sin(R_(2*Dm))
        -0.186*math.sin(R_(ms))-0.059*math.sin(R_(2*Mm-2*Dm))
        -0.057*math.sin(R_(Mm-2*Dm+ms))+0.053*math.sin(R_(Mm+2*Dm))
        +0.046*math.sin(R_(2*Dm-ms))+0.041*math.sin(R_(Mm-ms))
        -0.035*math.sin(R_(Dm))-0.031*math.sin(R_(Mm+ms))))
    eclLat=eclLat+(-0.173*math.sin(R_(f-2*Dm))-0.055*math.sin(R_(Mm-f-2*Dm))
        -0.046*math.sin(R_(Mm+f-2*Dm))+0.033*math.sin(R_(f+2*Dm))
        +0.017*math.sin(R_(2*Mm+f)))
    eps=23.4393-0.0000003563*d; er=R_(eps); lonr=R_(eclLon); latr=R_(eclLat)
    xe=math.cos(lonr)*math.cos(latr)
    yE=math.sin(lonr)*math.cos(latr)*math.cos(er)-math.sin(latr)*math.sin(er)
    ze=math.sin(lonr)*math.cos(latr)*math.sin(er)+math.sin(latr)*math.cos(er)
    ra=_norm(_atn2(yE,xe)); dec=_atn2(ze,math.hypot(xe,yE))
    nd=jd-2451545.0; gmst=_norm(280.46061837+360.98564736629*nd); lst=_norm(gmst+lng)
    ha=_norm(lst-ra)
    if ha>180: ha-=360
    az,alt=_radec_to_altaz(ha,dec,lat)
    mpar=math.degrees(math.asin(1.0/r))           # topocentric parallax in altitude
    alt=alt-mpar*math.cos(math.radians(alt))
    return az,alt

# ---- GC Arch ephemeris (mirrors Astro.bas GetGCArchPosition) ----
# Virtual point: the perpendicular to the line joining the two horizon "feet"
# of the Milky Way band = the AZIMUTH OF THE GALACTIC POLE (the pole of the
# b=0 great circle). Computed directly - no feet scan, no flip: the pole never
# nears the zenith here (max alt ~27 deg) so the bearing is smooth across the
# whole GC rise->set window. arch_rise = galactic NORTH pole azimuth (east at
# GC rise); arch_set = +180. Altitude 0 (horizon bearing; held pitch = Rp).
_GNP_RA=192.85948; _GNP_DEC=27.12825      # galactic north pole, J2000
def _gcarch_pole_az(local_dt,lat,lng,utc_off):
    utc=local_dt-datetime.timedelta(hours=utc_off)
    n=_date_to_julian(utc)-2451545.0
    gmst=_norm(280.46061837+360.98564736629*n); lst=_norm(gmst+lng)
    ha=_norm(lst-_GNP_RA)
    if ha>180: ha-=360
    z,a=_radec_to_altaz(ha,_GNP_DEC,lat)   # az, alt
    return z
def gcarch_rise_az_alt(local_dt,lat,lng,utc_off):
    """arch_rise points AT the arch at the start = galactic SOUTH pole az (NP+180); altitude 0."""
    return _norm(_gcarch_pole_az(local_dt,lat,lng,utc_off)+180.0),0.0
def gcarch_set_az_alt(local_dt,lat,lng,utc_off):
    """arch_set = the opposite perpendicular = galactic north pole az; altitude 0."""
    return _norm(_gcarch_pole_az(local_dt,lat,lng,utc_off)),0.0

EPHEM={"gc":gc_az_alt,"sun":sun_az_alt,"moon":moon_az_alt,
       "arch_rise":gcarch_rise_az_alt,"arch_set":gcarch_set_az_alt}
BAND_ALT_DEG=70.0       # zenith-band yaw ease threshold (matches AstroPush; mw/GC only)
def sample_track(obj,start_dt,dur_min,lat,lng,utc_off,step_min=0.5):
    """Sampled astro track over [start_dt, start_dt+dur] as the cart EXECUTES it.
    Yaw unwrapped continuously. For GC AND moon, samples above BAND_ALT_DEG are
    replaced by a smoothstep yaw ease (the azimuth-whip guard, matching
    AstroPush). Moon transits >70 deg ~1 week/month at this latitude, so it hits
    the same whip; sun never does, so sun has no zenith ease. Pitch clamped to 0
    below horizon (operator rule: below horizon -> use yaw, pitch=0 -> rides rim)."""
    fn=EPHEM.get(obj)
    if fn is None: return None
    n=max(2,int(round(dur_min/step_min))+1)
    ts=[start_dt+datetime.timedelta(minutes=i*dur_min/(n-1)) for i in range(n)]
    raw=[fn(t,lat,lng,utc_off) for t in ts]
    yaw=[a for a,_ in raw]; alt=[h for _,h in raw]
    for k in range(1,n):                        # global unwrap (continuous yaw frame)
        while yaw[k]-yaw[k-1]>180: yaw[k]-=360
        while yaw[k]-yaw[k-1]<-180: yaw[k]+=360
    if obj in ("gc","moon"):                    # zenith-band ease (GC/mw AND moon)
        eK=[k for k in range(n) if alt[k]>BAND_ALT_DEG]
        if len(eK)>=2:
            e0,e1=eK[0],eK[-1]; y0,y1=yaw[e0],yaw[e1]
            for k in range(e0,e1+1):
                f=(k-e0)/(e1-e0) if e1>e0 else 0.0
                yaw[k]=y0+(f*f*(3-2*f))*(y1-y0)
    return [(yaw[k]%360.0, max(0.0,alt[k]), alt[k]) for k in range(n)]   # (world, pitch_clamped, true_alt)

# ---------- Step A: read + resolve ----------
def resolve(path):
    wb=load_workbook(path,read_only=True,data_only=True)

    s={}
    for row in wb["Settings"].iter_rows(min_row=1,max_row=60,max_col=4,values_only=True):
        if row[1] is not None: s[str(row[1]).strip()]=row[2]
    lat=float(s.get("Latitude",0)); lon=float(s.get("Longitude",0))
    utc_off=float(s.get("UTC Offset (hrs)",9.5))
    # Track windows sample against the operator's planned start. The workbook
    # stores clock-of-day only (Shoot start anchor / col Q fire-times), no date,
    # so the renderer supplies the date = today. This mirrors the cubic author's
    # split: visualise the plan on the assumed day; the real press-time is the
    # cart's job via the realtime anchor, not the picture's.
    plan_date=datetime.date.today()
    # Night anchor (clock-of-day) for midnight rollover: a fire-time earlier than
    # this belongs to the NEXT calendar day (post-midnight part of the night).
    ssa=s.get("Shoot start anchor")
    shoot_min=(ssa.hour*60+ssa.minute) if isinstance(ssa,datetime.time) else 12*60

    arows=list(wb["AstroTable"].iter_rows(values_only=True))
    def cell(r,i):
        try: return float(r[i])
        except (TypeError,ValueError,IndexError): return None
    astro=[]
    for r in arows[1:]:
        if r[0] is None: continue
        gca,gcl,sa,sal=cell(r,1),cell(r,2),cell(r,3),cell(r,4)
        if None in (gca,gcl,sa,sal): continue
        rec=dict(t=str(r[0]),gc_az=gca,gc_alt=gcl,sun_az=sa,sun_alt=sal)
        ma,mal=cell(r,6),cell(r,7)            # cols G/H = Moon Az/Alt (present after step 3)
        if ma is not None and mal is not None:
            rec["moon_az"]=ma; rec["moon_alt"]=mal
        astro.append(rec)

    prows=list(wb["Plan"].iter_rows(min_row=1,max_row=80,max_col=30,values_only=True))
    # header row (the MIDDLE plan header carries "Step")
    hdr=next(i for i,r in enumerate(prows) if any(str(v).strip()=="Step" for v in r if v))

    # --- header-name -> column-index map (MIDDLE plan) -------------------
    # Read columns by HEADER NAME, not fixed letter, so the operator can
    # reorder MIDDLE columns in Excel without breaking this reader. Required
    # headers fail loud (clear error) rather than silently reading the wrong
    # column. Header strings are the contract; keep them stable in the sheet.
    hrow=prows[hdr]
    colmap={str(v).strip():i for i,v in enumerate(hrow) if v is not None and str(v).strip()}
    REQUIRED=["Step","Anchor ref","Fires at","Actual (mins)","Action","Target",
              "Ry (\u00b0)","Rp (\u00b0)","\u0394yaw (\u00b0)","\u0394pitch (\u00b0)","Dir (CW/CCW)"]
    missing=[h for h in REQUIRED if h not in colmap]
    if missing:
        raise KeyError("Plan MIDDLE header(s) not found: %s. Found: %s"
                       %(missing, sorted(colmap)))
    def C(name): return colmap[name]   # column index for a header

    # robust time-of-day in minutes: handles datetime.time/datetime objects
    # (what openpyxl returns) and "H:M" strings; date part is ignored.
    def _tmin(v):
        if v is None: return None
        if isinstance(v,(datetime.time,datetime.datetime)): return v.hour*60+v.minute
        try:
            pp=str(v).split(":"); return int(pp[0])*60+int(pp[1])
        except Exception: return None

    # cart WP -> heading (col B id, col H heading, col J Commences). wp_sched is
    # the ordered (commence_min, heading) list used to find the cart's expected
    # heading where it is PARKED at a GP's fire time -- the single heading rule
    # used by the push and validation (cart-position-at-fire-time, not anchor).
    wp_hdg={}; wp_sched=[]
    for r in prows[hdr+1:]:
        if r[1] and str(r[1]).strip().startswith("WP"):
            try: hd=float(r[7])
            except (TypeError,ValueError): continue
            wp_hdg[str(r[1]).strip()]=hd
            cm=_tmin(r[9])                         # col J Commences
            if cm is not None: wp_sched.append((cm,hd))
    wp_sched.sort()

    def cart_heading_at(fmin):
        # heading of the latest cart WP whose Commence <= fire time
        if not wp_sched: return 0.0
        if fmin is None: return wp_sched[0][1]
        h=wp_sched[0][1]
        for cm,hd in wp_sched:
            if cm<=fmin: h=hd
            else: break
        return h

    def fire_minutes(q):
        # Robust to the types openpyxl yields for a "Fires at" cell: a
        # datetime.time / datetime.datetime (data_only) OR an "HH:MM[:SS]"
        # string. The old str-split assumed a clean "HH:MM" and returned None
        # for datetime objects (date prefix) -> astro_az got None -> world 0.
        if q is None: return None
        if isinstance(q,(datetime.time,datetime.datetime)):
            return q.hour*60+q.minute
        try:
            h,m,*_=str(q).split(":"); return int(h)*60+int(m)
        except Exception: return None
    KEYS={"sun":("sun_az","sun_alt"),"gc":("gc_az","gc_alt"),"moon":("moon_az","moon_alt")}
    def astro_az(obj,fire_q):
        # LIVE ephemeris at the dated fire-time (single source of truth with the
        # cubic author and the track sampler), not a stale AstroTable lookup. The
        # AstroTable is 15-min, today-4pm-relative and only refreshed at Init
        # Shoot, so it misses/zeroes a GP whose fire-time or date differs (the
        # MW->GC rename + overnight plans). Computing live removes that whole
        # class: same gc_az_alt/sun_az_alt/moon_az_alt used everywhere else.
        fmin=fire_minutes(fire_q)
        if fmin is None: return None,None
        day=plan_date + (datetime.timedelta(days=1) if fmin<shoot_min-1 else datetime.timedelta(0))
        when=datetime.datetime.combine(day,datetime.time())+datetime.timedelta(minutes=fmin)
        fn={"sun":sun_az_alt,"gc":gc_az_alt,"moon":moon_az_alt,
            "arch_rise":gcarch_rise_az_alt,"arch_set":gcarch_set_az_alt}.get(obj)
        if fn is None: return None,None
        return fn(when,lat,lon,utc_off)

    gps=[]
    plan_end=None      # the gimbal-plan terminator: (step, "HH:MM") from the END row
    for r in prows[hdr+1:]:
        step=r[C("Step")]
        if step is None or str(step).strip()=="": continue
        act=str(r[C("Action")]).strip() if r[C("Action")] else ""
        if act.upper()=="END":
            et=r[C("Fires at")]                       # fire-time of the END row
            if isinstance(et,datetime.datetime):   ets="%02d:%02d"%(et.hour,et.minute)
            elif isinstance(et,datetime.time):     ets="%02d:%02d"%(et.hour,et.minute)
            else:                                  ets=str(et)[:5] if et else ""
            plan_end=(str(step).strip(),ets)
            break
        anchor=str(r[C("Anchor ref")]).strip() if r[C("Anchor ref")] else ""
        target=str(r[C("Target")]).strip().lower() if r[C("Target")] else ""
        target=target if target in ("sun","gc","moon","arch_rise","arch_set") else ""
        def num(c):
            try: return float(r[c])
            except (TypeError,ValueError): return None
        ry,rp,dy,dp=num(C("Ry (\u00b0)")),num(C("Rp (\u00b0)")),num(C("\u0394yaw (\u00b0)")),num(C("\u0394pitch (\u00b0)"))
        dlbl=str(r[C("Dir (CW/CCW)")]).strip().upper() if r[C("Dir (CW/CCW)")] else None
        dlbl=dlbl if dlbl in ("CW","CCW") else None

        # Track / Track-yaw on an astro target -> sample the full executed
        # track over the GP's window (fire-time .. + total dur) rather than a
        # single fire-time point. sun/moon/gc/arch_rise/arch_set; gc/moon carry
        # the zenith-band ease, arch is a smooth horizon-feet bearing (no ease).
        track=None
        ASTRO_TRACK_TARGETS=("sun","moon","gc","arch_rise","arch_set")
        if act.lower() in ("track","track-yaw") and target in ASTRO_TRACK_TARGETS:
            def _hms(v):
                if v is None: return None
                if isinstance(v,(datetime.time,datetime.datetime)): return v.hour*60+v.minute+v.second/60.0
                try: h,m,*sec=str(v).split(":"); return int(h)*60+int(m)+(int(sec[0]) if sec else 0)/60.0
                except Exception: return None
            fmin=_hms(r[C("Fires at")])
            _dv=r[C("Actual (mins)")]          # derived real window, minutes (number)
            dmin=float(_dv) if isinstance(_dv,(int,float)) else _hms(_dv)
            if fmin is not None and dmin:
                # roll past midnight: a fire-time before the night's start anchor
                # is on the following calendar day
                day=plan_date + (datetime.timedelta(days=1) if fmin<shoot_min-1 else datetime.timedelta(0))
                start_dt=datetime.datetime.combine(day,datetime.time())+datetime.timedelta(minutes=fmin)
                track=sample_track(target,start_dt,dmin,lat,lon,utc_off)
                # Track-yaw holds a FIXED pitch (Rp); the cubic's own pitch
                # (0 for arch - a horizon bearing) is not used. Re-pitch every
                # sample to Rp so the rendered sweep sits at the held elevation.
                if act.lower()=="track-yaw" and track and rp is not None:
                    track=[(w,rp,a) for (w,_p,a) in track]

        # NON-cumulative, per-GP. Ry/Rp = real-world (earth) anchor; blank Ry = cart-frame.
        if ry is not None:                       # earth-frame: Ry is the world bearing
            world=(ry+(dy or 0))%360
            frame="earth"; below=False; oaz=oalt=None
            pitch=(rp if rp is not None else 0.0)+(dp or 0.0)
        elif target:                              # earth-frame astro: object az at fire time
            oaz,oalt=astro_az(target,r[C("Fires at")])
            world=((oaz or 0)+(dy or 0))%360
            # arch_rise/arch_set are galactic-pole BEARINGS, valid all night - their
            # "altitude" is meaningless (returns ~0/<=0), so they must NEVER be
            # classified below-horizon (that draws a spurious goto-rise+wait marker
            # on a Move/Track to arch). Pitch comes from Rp, not altitude.
            is_arch = target in ("arch_rise","arch_set")
            frame="earth"; below=(False if is_arch else (oalt is not None and oalt<=0))
            pitch=(rp if rp is not None else 0.0)+(dp or 0.0)
        else:                                     # chassis-frame: offset from cart heading
            h=cart_heading_at(_tmin(r[C("Fires at")]))
            world=(h+(dy or 0))%360
            frame="chassis"; below=False; oaz=oalt=None
            pitch=(rp if rp is not None else 0.0)+(dp or 0.0)
        cart_hdg=cart_heading_at(_tmin(r[C("Fires at")]))   # cart-position heading
        gps.append(dict(step=str(step).strip(),act=act,frame=frame,target=target,
                        anchor=anchor,world=world,yaw_off=(dy or 0.0),ry=ry,pitch=pitch,
                        below=below,obj_az=oaz,obj_alt=oalt,dir=dlbl,track=track,
                        cart_hdg=cart_hdg))
    # cable wind-up: gimbal's running cart-frame angle (unwrapped), even though
    # plan yaw values are per-GP references. cart-frame yaw = world - heading.
    cable=0.0; prev_cf=None
    for g in gps:
        h=g["cart_hdg"]
        cf=((g["world"]-h+540)%360)-180          # cart-frame target, wrapped to +/-180
        if prev_cf is None:
            cable=cf; step=0.0
        else:
            short=((cf-prev_cf+540)%360)-180          # shortest signed step
            if g["dir"]=="CW":   step=short if short>=0 else short+360
            elif g["dir"]=="CCW":step=short if short<=0 else short-360
            else:                step=short            # auto = shortest
            cable+=step
        g["cf_yaw"]=cf; g["cable"]=cable; g["slew"]=step
        prev_cf=cf
    return dict(lat=lat,lon=lon,astro=astro,gps=gps,wp_hdg=wp_hdg,plan_end=plan_end)

# ---------- Step C: render ----------
def render(d,out,hi=None,map_img=None):
    fig,ax=plt.subplots(figsize=(9.5,9.5),dpi=160)
    fig.patch.set_facecolor(BG); ax.set_facecolor(BG)
    ax.set_xlim(-1.38,1.38); ax.set_ylim(-1.50,1.34); ax.set_aspect("equal"); ax.axis("off")

    if map_img:
        try:
            im=plt.imread(map_img); ax.imshow(im,extent=[-R,R,-R,R],zorder=0,alpha=0.5)
            clip=Circle((0,0),R,transform=ax.transData)
            for a in ax.images: a.set_clip_path(clip)
        except Exception as e: print("map underlay skipped:",e)

    for alt,lab in [(0,"horizon"),(30,"30\u00b0"),(60,"60\u00b0")]:
        rr=alt_to_r(alt); ax.add_patch(Circle((0,0),rr,fill=False,ec=RING,lw=1,zorder=1))
        if alt>0: ax.text(0,rr," "+lab,color=RINGTXT,fontsize=7,va="bottom",ha="left",zorder=2)
    ax.plot(0,0,marker="+",color=RINGTXT,ms=8,zorder=2)
    for az in range(0,360,30):
        x,y=az_to_xy(az,R); ax.plot([0,x],[0,y],color=RING,lw=0.6,zorder=1)
    for az,lab in [(0,"N"),(90,"E"),(180,"S"),(270,"W")]:
        x,y=az_to_xy(az,R*1.12); ax.text(x,y,lab,color=CARD,fontsize=14,fontweight="bold",
                                          ha="center",va="center",zorder=3)

    # faint astro context arcs (sun, gc, + moon when the table has it)
    arc_specs=[(("sun_az","sun_alt"),EARTH["sun"]),(("gc_az","gc_alt"),EARTH["gc"])]
    if any("moon_az" in p for p in d["astro"]):
        arc_specs.append((("moon_az","moon_alt"),EARTH["moon"]))
    for key,col in arc_specs:
        seg=[az_to_xy(p[key[0]],alt_to_r(p[key[1]])) for p in d["astro"]
             if key[0] in p and p[key[1]]>0]
        if seg: ax.plot([p[0] for p in seg],[p[1] for p in seg],color=col,lw=1.2,alpha=0.30,zorder=2)

    gps=d["gps"]

    # ---- executed astro tracks: swept polyline over the GP window ----
    # (world az -> angle, clamped pitch -> radius). Below-horizon samples ride
    # the rim (pitch clamped to 0); the rise crossing gets a tick.
    pe=d.get("plan_end")
    track_gps=[g for g in gps if g.get("track")]
    last_track=track_gps[-1] if track_gps else None
    for gp in gps:
        tr=gp.get("track")
        if not tr: continue
        col=EARTH.get(gp["target"],EARTH_DEF)
        # densify for drawing only: interpolate world az + clamped pitch between
        # samples so fast-motion sections (low-alt rise) read as a smooth curve,
        # not a sample-to-sample chord. Does not change the executed-frame data.
        SUB=8
        dpts=[]
        for k in range(len(tr)-1):
            w0,p0,_=tr[k]; w1,p1,_=tr[k+1]
            dw=((w1-w0+540)%360)-180                 # shortest arc between samples
            for j in range(SUB):
                f=j/SUB
                dpts.append(az_to_xy((w0+dw*f)%360, alt_to_r(p0+(p1-p0)*f)))
        dpts.append(az_to_xy(tr[-1][0],alt_to_r(tr[-1][1])))
        pts=[az_to_xy(w,alt_to_r(p)) for (w,p,_a) in tr]   # sample points (for rise/badge/arrow)
        ax.plot([q[0] for q in dpts],[q[1] for q in dpts],color=col,lw=2.4,
                alpha=0.95,zorder=5,solid_capstyle="round")
        # rise crossing (first sample with true altitude >= 0)
        for k in range(1,len(tr)):
            if tr[k-1][2]<0<=tr[k][2]:
                rx,ry=pts[k]
                ax.plot(rx,ry,marker="|",color=col,ms=12,mew=2,zorder=6)
                # label radially OUTWARD (off the rim) so it clears the start badge
                lx,ly=az_to_xy(tr[k][0],R*1.06)
                ax.annotate("rise",(lx,ly),color=col,fontsize=7,ha="center",va="center",zorder=8)
                break
        # direction arrowhead at the end of the sweep
        if len(pts)>=2:
            ax.add_patch(FancyArrowPatch(pts[-2],pts[-1],color=col,lw=0,arrowstyle="-|>",
                         mutation_scale=15,alpha=0.95,zorder=6))
        # gimbal-plan end: label the LAST track's endpoint with the END row (any GP#)
        if gp is last_track and pe:
            ex,ey=pts[-1]
            etxt=f"end {pe[0]}" + (f"  {pe[1]}" if pe[1] else "")
            ax.annotate(etxt,(ex,ey),color=col,fontsize=7.5,fontweight="bold",
                        ha="left",va="bottom",xytext=(7,5),textcoords="offset points",zorder=8)


    def leg_alpha(i):
        if hi is None: return 0.9
        names={gps[i]["step"],gps[i+1]["step"]}
        focus = hi if isinstance(hi,str) else "GP%02d"%hi
        return 0.9 if focus in names else DIM
    nlegs=len(gps)-1
    for i in range(nlegs):
        # Track-aware endpoints (defined below) decide both the skip and the draw.
        tr_i = gps[i].get("track")
        tr_j = gps[i+1].get("track")
        a0 = tr_i[-1][0] if tr_i else gps[i]["world"]    # leg starts where GP i LEAVES the gimbal
        a1 = tr_j[0][0]  if tr_j else gps[i+1]["world"]  # leg ends where GP i+1 PICKS UP
        # Skip a connector whose real travel is ~0 (e.g. a track that ends where
        # the next picks up) - the polylines already convey the motion.
        if abs(((a1-a0+540)%360)-180) < 2.0 and (tr_i or tr_j):
            continue
        legr=R*(0.72+0.055*i)                      # fan consecutive legs onto separate radii
        dlbl=gps[i+1]["dir"]
        short=((a1-a0+540)%360)-180
        if dlbl=="CW":    diff=short if short>=0 else short+360
        elif dlbl=="CCW": diff=short if short<=0 else short-360
        else:             diff=short               # auto = shortest
        amb = (dlbl is None) and abs(abs(short)-180)<=30   # only flag if not yet chosen
        n=max(2,int(abs(diff)/4))
        pts=[az_to_xy((a0+diff*k/n)%360,legr) for k in range(n+1)]
        la=leg_alpha(i)
        col_leg = MID if (dlbl and abs(diff)>180.5) else CARD   # highlight a forced long-way leg
        ax.plot([p[0] for p in pts],[p[1] for p in pts],color=col_leg,lw=1.6,alpha=0.6*la,
                ls=("--" if amb else "-"),zorder=4)
        ax.add_patch(FancyArrowPatch(pts[-2],pts[-1],color=col_leg,lw=0,arrowstyle="-|>",
                     mutation_scale=15,alpha=0.85*la,zorder=4))
        # leg label: number + chosen direction (or AUTO)
        sx,sy=az_to_xy((a0+diff*0.12)%360,legr)
        tag=f"{i+1}\u2192{i+2}" + (f" {dlbl}" if dlbl else "")
        ax.text(sx,sy,tag,color=col_leg,fontsize=7,alpha=0.8*la,ha="center",va="center",zorder=5)
        if amb:
            mx,my=az_to_xy((a0+short/2)%360,legr)
            ax.plot(mx,my,"D",color=MID,ms=8,alpha=la,zorder=5,mec=BG,mew=1)
            ax.annotate(f"leg {i+1}\u2192{i+2}: ~180\u00b0 \u2014 set CW/CCW (col AC)",(mx,my),
                        color=MID,fontsize=7,alpha=la,ha="center",va="bottom",
                        xytext=(0,9+10*i),textcoords="offset points",zorder=8)

    prev=None; seen={}; idx=1
    for gp in gps:
        foreground = (hi is None) or (gp["step"]==hi) or (gp["step"]==("GP%02d"%hi) if isinstance(hi,int) else False)
        a = 1.0 if foreground else DIM
        col = CHASSIS if gp["frame"]=="chassis" else EARTH.get(gp["target"],EARTH_DEF)

        # GP with a sampled track: the swept polyline already shows it. Badge the
        # start of the sweep and move on (no single-point glyph / park marker).
        tr=gp.get("track")
        if tr:
            sx,sy=az_to_xy(tr[0][0],alt_to_r(tr[0][1]))
            ax.plot(sx,sy,"o",color=col,ms=15,alpha=a,zorder=8,mec=BG,mew=1.5)
            ax.text(sx,sy,str(idx),color=BG,fontsize=9,fontweight="bold",ha="center",va="center",
                    alpha=a,zorder=9)
            # label sits just OUTSIDE the rim at the start azimuth (start is on the
            # rim when below-horizon), so it never overlaps the badge or rise tick
            lx,ly=az_to_xy(tr[0][0],R*1.13)
            ax.annotate(f"{gp['step']} track {tr[0][0]:.0f}\u00b0\u2192{tr[-1][0]:.0f}\u00b0",
                        (lx,ly),color=col,fontsize=7.5,alpha=a,ha="center",va="center",zorder=8)
            prev=gp; idx+=1; continue

        # below-horizon earth object -> park-and-wait marker at rise, no glyph line
        if gp["frame"]=="earth" and gp["below"]:
            wx,wy=az_to_xy(gp["world"],R)
            ax.plot(wx,wy,marker="P",color=col,ms=14,alpha=a,zorder=6,mec=BG,mew=1)
            ax.text(wx,wy,str(idx),color=BG,fontsize=8,fontweight="bold",ha="center",va="center",
                    alpha=a,zorder=9)
            ax.annotate(gp["step"]+" goto-rise + wait",(wx,wy),color=col,fontsize=7.5,alpha=a,
                        ha="center",va="top",xytext=(0,-10),textcoords="offset points",zorder=8)
            prev=gp; idx+=1; continue

        # glyph: radial line rim -> endpoint at altitude(pitch); length ∝ pitch
        pitch=gp["pitch"]; flag = pitch>PITCH_LIMIT
        endr=alt_to_r(pitch)                         # near centre = high pitch
        ex,ey=az_to_xy(gp["world"],endr)
        rx,ry=az_to_xy(gp["world"],R)                # rim foot
        ls = "-" if gp["frame"]=="chassis" else (0,(6,2))
        ax.plot([rx,ex],[ry,ey],color=col,lw=2.6,ls=ls,alpha=a,zorder=5,solid_capstyle="round")
        ax.plot(ex,ey,"o",color=col,ms=8,alpha=a,zorder=6,mec=BG,mew=1)
        if flag:
            ax.plot(ex,ey,"o",mfc="none",mec=FLAG,ms=15,mew=2.2,alpha=a,zorder=7)
        lab=f"{gp['step']}  {pitch:.0f}\u00b0p"
        key=(round(gp["world"]),round(min(pitch,90)))
        dy=-12*seen.get(key,0); seen[key]=seen.get(key,0)+1
        ax.annotate(lab,(ex,ey),color=(FLAG if flag else col),fontsize=8,alpha=a,
                    ha="left",va="center",xytext=(8,dy),textcoords="offset points",zorder=8)

        # sequence badge on the glyph endpoint
        bx,by=az_to_xy(gp["world"],endr)
        ax.plot(bx,by,"o",color=col,ms=15,alpha=a,zorder=8,mec=BG,mew=1.5)
        ax.text(bx,by,str(idx),color=BG,fontsize=9,fontweight="bold",ha="center",va="center",
                alpha=a,zorder=9)
        prev=gp; idx+=1

    # legend
    handles=[
        Line2D([0],[0],color=CHASSIS,lw=2.6,label="chassis-frame GP (cart-nose + offset)"),
        Line2D([0],[0],color=EARTH_DEF,lw=2.6,ls=(0,(6,2)),label="earth-frame GP (points at object)"),
        Line2D([0],[0],color=CARD,lw=1.6,marker=">",alpha=0.7,label="GP sweep order (1\u21922\u21923\u21924)"),
        Line2D([0],[0],color=MID,marker="D",lw=1,ls="--",label="~180\u00b0 sweep \u2014 direction ambiguous, set midpoint"),
        Line2D([0],[0],marker="o",mfc="none",mec=FLAG,lw=0,ms=10,label="pitch > 80\u00b0 flag"),
        Line2D([0],[0],color=EARTH["gc"],lw=1.2,alpha=0.4,label="astro context arcs (sun / GC)"),
        Line2D([0],[0],color=EARTH["gc"],lw=2.4,label="astro track (executed sweep; rim = below horizon)"),
    ]
    ax.legend(handles=handles,loc="lower center",bbox_to_anchor=(0.5,-0.11),ncol=1,
              frameon=False,fontsize=8.5,labelcolor=TXT,handlelength=2.4)
    sub = "" if hi is None else f"   |   focus: {hi if isinstance(hi,str) else 'GP%02d'%hi}"
    ax.set_title("Gimbal Plan View  —  glyph dir = world bearing, length = pitch"+sub,
                 color=TXT,fontsize=12,pad=14)
    ax.text(0,-1.40,f"lat {d['lat']:.4f}, lon {d['lon']:.4f}   |   cart at centre, N up, radius = altitude",
            color=RINGTXT,fontsize=8,ha="center")

    fig.savefig(out,facecolor=BG,bbox_inches="tight")
    svg=out.rsplit(".",1)[0]+".svg"; fig.savefig(svg,facecolor=BG,bbox_inches="tight")
    print("wrote",out,"and",svg)

if __name__=="__main__":
    ap=argparse.ArgumentParser()
    ap.add_argument("xlsm"); ap.add_argument("out",nargs="?",default="gimbal_planview_v2.png")
    ap.add_argument("--gp",type=int,default=None); ap.add_argument("--map",default=None)
    a=ap.parse_args()
    d=resolve(a.xlsm)
    for g in d["gps"]:
        print(g["step"],g["frame"],"world=%.0f"%g["world"],"pitch=%.0f"%g["pitch"],
              "cf_yaw=%+.0f"%g["cf_yaw"],"cable=%+.0f"%g["cable"],"slew=%+.0f"%g["slew"])
    render(d,a.out,hi=a.gp,map_img=a.map)
