#!/usr/bin/env python3
"""
HyperLapse — Gimbal Cable Strip (view #3).

Shows the gimbal's yaw sweep as an UNWRAPPED world bearing (relative to
true North), accumulated leg-by-leg in the direction the operator set in
Plan col AC (CW/CCW; blank = shortest). Plotted on a 1-D strip against the
450 deg span limit: min yaw on the left, min+450 (the hard ceiling where
cables break) on the right.

It reuses resolve() from gimbal_planview_v2 (single source of truth), so
frame/world/dir come out identical to the dial. This view adds only the
unwrap-to-world and the strip drawing.

The tool SHOWS what the plan says; it does not pick direction. Operator
sets col AC (proposed by the sweep-dir macro, accepted/overridden by the
operator using the dial and/or this strip).

Usage:
  python3 gimbal_cablestrip.py HyperLapse.xlsm out.png [--gp N] [--limit 450]
"""
import argparse
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import FancyArrowPatch, Rectangle
from gimbal_planview_v2 import (resolve, BG, RING, RINGTXT, CARD, CHASSIS,
                                MID, FLAG, TXT)

USED="#3b82f6"      # used-span fill
FREE="#1f2a37"      # headroom fill
LIMIT="#ff5470"     # span-limit line
PANO="#d9a441"      # pano-swing reach band (PanoCycle ±outer offset)

def portrait_pano_reach(path):
    """Max abs portrait pano offset (deg) from the PANO sheet, e.g. 89 for
    {-89,-45,0,45,89}. This is the half-swing the camera reaches at photos 1
    and X of every PanoCycle, on TOP of the tracked arch centre. Read from the
    sheet (like pano_planner) so it tracks the config, not hardcoded. Returns
    0.0 if not found (then no widening - silent, not wrong)."""
    try:
        from openpyxl import load_workbook
        wb=load_workbook(path,read_only=True,data_only=True)
        vals=[]
        for nm in ("panoP_offsets","panoP_o0","panoP_o1","panoP_o2","panoP_o3","panoP_o4"):
            try:
                dn=wb.defined_names[nm]
                for sheet,ref in dn.destinations:
                    ws=wb[sheet]
                    for row in ws[ref.replace("$","")]:
                        for c in row:
                            if isinstance(c.value,(int,float)): vals.append(abs(float(c.value)))
            except Exception: pass
        return max(vals) if vals else 0.0
    except Exception:
        return 0.0

def unwrap_cart(gps, wp_hdg):
    """Accumulate CART-FRAME yaw leg-by-leg honouring col AC.

    Cable wind is the gimbal's yaw RELATIVE TO THE CART BODY, not the world
    bearing -- the cables tangle around the cart. So we unwrap cart-frame yaw
    cf = world - anchor_heading, for both point GPs and every track sample.

    Returns (wu, sweeps): wu[i] is the unwrapped value AT each GP (track GPs use
    their END, since that's where the wind sits when the next leg begins);
    sweeps[i] is (u_start,u_end) for a track GP (the continuous yaw excursion
    over its window) or None for a point GP. Track samples unwrap continuously
    so a multi-turn sweep counts every degree toward the span."""
    wu=[]; sweeps=[]; prev_w=None; prev_u=0.0
    def step_dir(short,d):
        if d=="CW":    return short if short>=0 else short+360
        if d=="CCW":   return short if short<=0 else short-360
        return short                                  # auto = shortest
    for g in gps:
        h=g["cart_hdg"]                               # cart heading where parked at fire time
        tr=g.get("track")
        if tr:                                        # walk the sampled azimuths, cart-frame
            azs=[((s[0]-h+540)%360)-180 for s in tr]  # world -> cart-frame, wrapped
            if prev_w is None:
                u0=azs[0]
            else:
                # ENTRY into a track uses the SHORTEST path to its first sample.
                # Do NOT apply the row's CW/CCW here: a track row's dir describes
                # the *track* motion (handled by the per-sample unwrap below), not
                # the entry slew. Forcing dir on a near-zero entry delta (e.g. when
                # the previous Move already arrived at the track's start) injects a
                # phantom +/-360 turn - the cause of false cable-limit overruns.
                short=((azs[0]-prev_w+540)%360)-180
                u0=prev_u+short
            u=u0; pw=azs[0]
            for w in azs[1:]:
                short=((w-pw+540)%360)-180            # continuous unwrap within track
                u+=short; pw=w
            wu.append(u); sweeps.append((u0,u))
            prev_w=azs[-1]; prev_u=u
        else:
            w=((g["world"]-h+540)%360)-180            # cart-frame for the point GP
            if prev_w is None:
                u=w
            else:
                short=((w-prev_w+540)%360)-180
                u=prev_u+step_dir(short,g["dir"])
            wu.append(u); sweeps.append(None)
            prev_w=w; prev_u=u
    return wu,sweeps

def render(d, out, hi=None, limit=450.0, pano_reach=0.0):
    gps=d["gps"]
    if not gps:
        print("no GPs in plan"); return
    wu,sweeps=unwrap_cart(gps, d["wp_hdg"])
    # an acquire-length track (negligible excursion) shouldn't draw its own bar;
    # fold it out so only the real tracking sweep shows
    for i,sw in enumerate(sweeps):
        if sw and abs(sw[1]-sw[0])<1.0: sweeps[i]=None
    # PanoCycle widening: an arch-target track GP runs a PanoCycle every cadence
    # frame - the camera swings +/-reach (portrait outer offset, e.g. 89deg)
    # around the tracked centre at photos 1 and X. That swing is REAL cable wind
    # the operator must see, on TOP of the centre track. Compute a widened band
    # per arch sweep: (min-reach, max+reach). reach<=0 (no portrait cfg) = no-op.
    pano_band=[None]*len(sweeps)
    if pano_reach>0:
        for i,sw in enumerate(sweeps):
            if not sw: continue
            if gps[i].get("target") in ("arch_rise","arch_set"):
                a,b=sorted(sw)
                pano_band[i]=(a-pano_reach, b+pano_reach)
    # span must include the full excursion of any track sweep AND any pano band
    allvals=list(wu)
    for sw in sweeps:
        if sw: allvals+=[sw[0],sw[1]]
    for pb in pano_band:
        if pb: allvals+=[pb[0],pb[1]]
    lo=min(allvals); hi_used=max(allvals); span=hi_used-lo
    ceil=lo+limit; headroom=ceil-hi_used
    # which GP carries the max-wind point (point value or either sweep end)
    def maxwind_label():
        for i,(g,x) in enumerate(zip(gps,wu)):
            if abs(x-hi_used)<1e-6: return g["step"]
            sw=sweeps[i]
            if sw and (abs(sw[0]-hi_used)<1e-6 or abs(sw[1]-hi_used)<1e-6): return g["step"]
        return gps[0]["step"]

    fig,ax=plt.subplots(figsize=(11,3.4))
    ax.set_xlim(lo-15, ceil+15); ax.set_ylim(-1.0,1.4)
    ax.axis("off")

    y=0.0
    # full track, used span, headroom
    ax.add_patch(Rectangle((lo,y-0.13),ceil-lo,0.26,facecolor=FREE,edgecolor=RING,lw=1,zorder=1))
    ax.add_patch(Rectangle((lo,y-0.13),span,0.26,facecolor=USED,edgecolor="none",alpha=0.45,zorder=2))
    # limit line (right end) + min mark
    ax.plot([ceil,ceil],[y-0.32,y+0.32],color=LIMIT,lw=2.5,zorder=5)
    ax.text(ceil,y+0.42,f"span limit  {limit:.0f}\u00b0\n(cables break beyond)",
            color=LIMIT,fontsize=8,ha="center",va="bottom")
    ax.plot([lo,lo],[y-0.32,y+0.32],color=RINGTXT,lw=1.5,zorder=5)
    ax.text(lo,y-0.42,f"min  {lo:.0f}\u00b0",color=RINGTXT,fontsize=8,ha="center",va="top")

    # track sweep bars: the continuous yaw excursion during a Track window,
    # drawn as a bold segment on the strip (this is the wind the cables see)
    pe=d.get("plan_end")
    last_tr_i=max([i for i,sw in enumerate(sweeps) if sw], default=None)
    # PanoCycle reach bands (drawn FIRST, behind the centre track bars): the
    # +/-reach swing the camera makes at photos 1 and X of every cycle. Shows
    # the operator the FULL cable wind a PanoCycle GP produces, not just centre.
    for i,pb in enumerate(pano_band):
        if not pb: continue
        pa,pbb=sorted(pb)
        ax.add_patch(Rectangle((pa,y-0.13),pbb-pa,0.26,facecolor=PANO,edgecolor="none",
                     alpha=0.22,zorder=2.2))
        # label well BELOW the band, anchored at its LEFT edge, so it can't collide
        # with the centre-track label (above) or the END marker (just under band).
        ax.plot([pa,pa],[y-0.13,y-0.56],color=PANO,lw=0.8,alpha=0.7,zorder=5)
        ax.text(pa,y-0.60,f"{gps[i]['step']} PanoCycle reach \u00b1{pano_reach:.0f}\u00b0",
                color=PANO,fontsize=7.5,ha="left",va="top",zorder=6)
    for i,sw in enumerate(sweeps):
        if not sw: continue
        a,b=sorted(sw)
        ax.add_patch(Rectangle((a,y-0.08),b-a,0.16,facecolor=MID,edgecolor="none",
                     alpha=0.85,zorder=3))
        ax.add_patch(FancyArrowPatch((sw[0],y+0.20),(sw[1],y+0.20),
            connectionstyle="arc3,rad=0.0",arrowstyle="-|>",mutation_scale=12,
            color=MID,lw=1.6,alpha=0.9,zorder=4))
        ax.text((a+b)/2,y+0.30,f"{gps[i]['step']} track {sw[0]:.0f}\u00b0\u2192{sw[1]:.0f}\u00b0"
                f"  ({abs(b-a):.0f}\u00b0)",color=MID,fontsize=8,ha="center",va="bottom",zorder=6)
        # gimbal-plan end at the last track's end edge (END row step+time, any GP#)
        if i==last_tr_i and pe:
            etxt=f"end {pe[0]}" + (f"  {pe[1]}" if pe[1] else "")
            ax.plot([sw[1],sw[1]],[y-0.20,y+0.12],color=MID,lw=1.5,zorder=5)
            ax.text(sw[1],y-0.30,etxt,color=MID,fontsize=8,fontweight="bold",
                    ha="center",va="top",zorder=6)

    # sweep order arrows between consecutive GP-resting positions (point GPs only;
    # a track GP's own arrow is drawn above as the sweep)
    for i in range(len(wu)-1):
        if sweeps[i] or sweeps[i+1]: continue
        a,b=wu[i],wu[i+1]
        col=CHASSIS if b>=a else FLAG
        ax.add_patch(FancyArrowPatch((a,y+0.16),(b,y+0.16),
            connectionstyle="arc3,rad=0.45",arrowstyle="-|>",mutation_scale=12,
            color=col,lw=1.4,alpha=0.85,zorder=4))

    # alternate label side by x-order so near-coincident GPs don't overlap
    side={}; order=sorted(range(len(wu)),key=lambda i:wu[i])
    for k,i in enumerate(order): side[i]=(k%2==0)   # True=above

    # markers: point GPs get a dot+label; track GPs already have the sweep bar+label
    for i,(g,x) in enumerate(zip(gps,wu)):
        if sweeps[i]: continue
        # suppress a point GP that sits on the next GP's track start (acquire row)
        if i+1<len(sweeps) and sweeps[i+1] and abs(x-sweeps[i+1][0])<2.0: continue
        focus = (hi is not None and (g["step"]==(hi if isinstance(hi,str) else None)
                                     or i+1==hi))
        ismax=(abs(x-hi_used)<1e-6)
        c = MID if focus else (LIMIT if ismax else CARD)
        ax.plot([x],[y],marker="o",ms=11 if (focus or ismax) else 8,
                color=c,zorder=6,markeredgecolor=BG,markeredgewidth=1.2)
        lab=f"{g['step']}\n{wu[i]:.0f}\u00b0"
        if ismax: lab+="  (max)"
        up=side[i]
        ax.text(x, y+0.32 if up else y-0.32, lab, color=c,fontsize=8.5,
                ha="center", va="bottom" if up else "top", zorder=6)

    ax.set_title("Gimbal Cable Strip  —  cart-frame yaw wind (cable tangle, unwrapped via col AC) vs span limit",
                 color=TXT,fontsize=12,pad=12)
    sub=(f"used span {span:.0f}\u00b0   |   headroom to limit {headroom:.0f}\u00b0   |   "
         f"max-wind {maxwind_label()} at {hi_used:.0f}\u00b0")
    ax.text((lo+ceil)/2,-0.78,sub,color=(LIMIT if headroom<0 else RINGTXT),
            fontsize=9,ha="center")

    # over-limit: unmissable banner + red wash over the strip. Cables would break;
    # Prep Cart will refuse to push this plan.
    if headroom<0:
        ax.axhspan(y-0.13,y+0.13,xmin=0,xmax=1,facecolor=LIMIT,alpha=0.18,zorder=2.5)
        ax.text((lo+ceil)/2, 1.18,
                f"\u26a0  EXCEEDS {limit:.0f}\u00b0 CABLE LIMIT  by {-headroom:.0f}\u00b0  \u2014  CART PUSH BLOCKED",
                color="#ffffff",fontsize=12,fontweight="bold",ha="center",va="center",zorder=10,
                bbox=dict(boxstyle="round,pad=0.5",facecolor=LIMIT,edgecolor="none"))

    fig.savefig(out,facecolor=BG,bbox_inches="tight")
    svg=out.rsplit(".",1)[0]+".svg"; fig.savefig(svg,facecolor=BG,bbox_inches="tight")
    print("wrote",out,"and",svg)
    print("min %.0f  max %.0f  span %.0f  headroom %.0f"%(lo,hi_used,span,headroom))
    # sidecar for the VBA cable-span guard (single source of truth):
    # one line 'span headroom limit' so Prep Cart can gate without recomputing.
    try:
        import os
        side=os.path.join(os.path.dirname(os.path.abspath(out)),"cablestrip_span.txt")
        open(side,"w").write("%.0f %.0f %.0f\n"%(span,headroom,limit))
    except Exception as e:
        print("sidecar write failed:",e)
    # per-GP file for the cart's on-screen cable strip (CableStripPush.bas). The
    # renderer already has every GP's unwrapped cart-frame wind; this writes it
    # out so the cart strip can draw the TRACK sweeps it currently skips, WITHOUT
    # recomputing (single source). Strip x is 0..STRIP_W mapped over [lo, ceil],
    # so VBA just places what it reads. One line per GP:
    #   point GP : "P <step> <x>"
    #   track GP : "T <step> <x_start> <x_end>"
    # Header line first: "STRIP <STRIP_W> <lo> <ceil>".
    try:
        import os
        STRIP_W=355.0                              # cart strip width (matches chart contract)
        rng=(ceil-lo) if (ceil-lo)>1e-6 else 1.0
        def to_x(u): return (u-lo)/rng*STRIP_W
        gp_file=os.path.join(os.path.dirname(os.path.abspath(out)),"cablestrip_gps.txt")
        with open(gp_file,"w") as fh:
            fh.write("STRIP %.0f %.2f %.2f\n"%(STRIP_W,lo,ceil))
            for i,g in enumerate(gps):
                sw=sweeps[i]
                if sw:
                    # arch (PanoCycle) GP: emit the WIDENED band so the cart strip
                    # shows the +/-reach swing too; else the bare centre sweep.
                    pb=pano_band[i]
                    if pb:
                        fh.write("T %s %.2f %.2f\n"%(g["step"],to_x(pb[0]),to_x(pb[1])))
                    else:
                        fh.write("T %s %.2f %.2f\n"%(g["step"],to_x(sw[0]),to_x(sw[1])))
                else:
                    fh.write("P %s %.2f\n"%(g["step"],to_x(wu[i])))
    except Exception as e:
        print("per-GP file write failed:",e)

if __name__=="__main__":
    ap=argparse.ArgumentParser()
    ap.add_argument("xlsm"); ap.add_argument("out",nargs="?",default="gimbal_cablestrip.png")
    ap.add_argument("--gp",type=int,default=None)
    ap.add_argument("--limit",type=float,default=450.0)
    a=ap.parse_args()
    d=resolve(a.xlsm)
    reach=portrait_pano_reach(a.xlsm)
    render(d,a.out,hi=a.gp,limit=a.limit,pano_reach=reach)
